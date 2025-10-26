import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import functions
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import glob
import os
import pandas as pd

def browse_folder():
    folder_path = filedialog.askdirectory(title="Select Folder Containing CSV Files")
    if folder_path:
        folder_var.set(folder_path)
        folder_entry.xview_moveto(1)  # Scroll to the end of the entry

def browse_file(index):
    file_path = filedialog.askopenfilename(
        title="Select a CSV file",
        filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
    
    if file_path:
        entry_vars[index].set(file_path)
        file_entries[index].xview_moveto(1)  # Scroll to the end of the entry

def toggle_mode():
    folder_enabled = mode_var.get() == "folder"

    folder_entry.config(state="normal" if folder_enabled else "disabled")
    folder_browse.config(state="normal" if folder_enabled else "disabled")
    
    for entry, button in zip(file_entries, file_buttons):
        state = "disabled" if folder_enabled else "normal"
        entry.config(state=state)
        button.config(state=state)

def browse_output_folder():
    output_path = filedialog.askdirectory(title="Select Output Folder")
    if output_path:
        output_folder_var.set(output_path)
        output_folder_entry.xview_moveto(1)

def create_merged_file():
    filename = output_filename_var.get().strip()
    folder = output_folder_var.get().strip()

    if not filename or not folder:
        messagebox.showerror("Error", "Please specify both filename and folder/files.")
        return
    
    file_exist = os.path.isfile(os.path.join(folder, filename + ".xlsx"))
    if file_exist:
        result = messagebox.askokcancel("File Exists", f"The file already exists in the selected folder. Do you want to overwrite it?")
        if not result:
            return
    functions.create_by_merging(get_csv_files(), folder, filename)
    messagebox.showinfo("Success", f"File will be saved as:\n{folder}/{filename}.xlsx.. \n\nLoading the headers...")

    # auto load the header files
    load_headers_from_file(folder, filename)

def update_merged_file():
    csv_files = get_csv_files()
    column_id = id_var.get().strip()
    filename = output_filename_var.get().strip()
    folder = output_folder_var.get().strip()
    merged_file = os.path.join(folder, filename + ".xlsx")
    replace_n = 0

    if not os.path.exists(merged_file):
        messagebox.showerror("Error", "Merged file does not exist. Please create it first.")
        return

    # Read the current merged data
    current_df = pd.read_excel(merged_file, sheet_name="Raw Data")
    for csv_file in csv_files:
        new_df = pd.read_csv(csv_file)
        # Ensure column exists
        if column_id not in new_df.columns:
            messagebox.showerror("Error", f"Column '{column_id}' not found in {os.path.basename(csv_file)}")
            continue

        # CSV should have only 1 unique ID
        csv_ids = new_df[column_id].unique()

        if len(csv_ids) != 1:
            messagebox.showwarning("Warning", f"{os.path.basename(csv_file)} contains multiple {column_id} values. Using first one.")
        
        csv_id = str(csv_ids[0])

        # Check if ID exists already in merged file
        if column_id in current_df.columns and (current_df[column_id].astype(str) == csv_id).any():
            # Ask user whether to replace
            response = messagebox.askyesno(
                "ID Already Exists",
                f"'{csv_id}' already exists in merged data.\n\nReplace it?"
            )
            if not response:
                continue
            else:
                replace_n += 1

            current_df = current_df[current_df[column_id].astype(str) != csv_id] # Replace → remove old rows for this ID

        current_df = pd.concat([current_df, new_df], ignore_index=True) # Append new data

    if replace_n == 0:
        messagebox.showinfo("Info", "No matching IDs were found to update. File remains unchanged.")
        return 

    # Save updated merged data
    wb = load_workbook(merged_file)
    ws = wb["Raw Data"] 
    
    # clear old data except header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    
    wb.save(merged_file)

    # Append updated data
    with pd.ExcelWriter(merged_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        writer._book = wb
        writer._sheets = {ws.title: ws for ws in wb.worksheets}
        current_df.to_excel(writer, sheet_name="Raw Data", index=False, startrow=1, header=False)

    # Convert the Raw Data to Table to easily refresh pivot chart later
    functions.convert_to_RawData_table(merged_file)

    # rebind and show buttons
    functions.rebindPivotSources(merged_file)

    messagebox.showinfo("Success", f"Updated merged file. \n\nReplaced {replace_n} entries based on '{column_id}'. \n\nLoading headers...")
    # auto load the header files
    load_headers_from_file(folder, filename)

def generate_pivot_chart():
    pc_axis = axis_var.get().strip()
    pc_legend = legend_var.get().strip()
    pc_values = values_var.get().strip()
    pivotName = pivotChart_var.get().strip()
    filename = output_filename_var.get().strip()
    folder = output_folder_var.get().strip()
    merged_file = os.path.join(folder, filename + ".xlsx")
    functions.generate_pivotChart(merged_file, pc_axis, pc_legend, pc_values, pivotName)
    
def get_csv_files():
    if mode_var.get() == "folder":
        return glob.glob(os.path.join(folder_var.get().strip(), "*.csv")) # returns list of all csv file path
    return [var.get().strip() for var in entry_vars if var.get().strip() != ""] 

def load_headers_from_file(folder, filename):
    merged_file = os.path.join(folder, filename + ".xlsx")
    df = pd.read_excel(merged_file, sheet_name="Raw Data", nrows=1)  # just read the header row
    print("df header columns:", df.columns.tolist())
    header_listbox.delete(0, tk.END)
    for col in df.columns:
        header_listbox.insert(tk.END, col)
    print("Done printing headers to listbox.")

# Main window
root = tk.Tk()
root.geometry("1100x700")
root.resizable(False, False)
root.title("Multiple CSVs to Pivot Chart")

# LEFT FRAME
left_frame = tk.Frame(root, width=700, height=650)
left_frame.pack_propagate(False) # False - use the defined size
left_frame.pack(side="left", padx=10, pady=10)

# Frame for mode
mode_frame = ttk.Frame(left_frame, width=700, height=50, borderwidth=10, relief=tk.GROOVE)
mode_frame.pack_propagate(False) # False - use the defined size
mode_frame.pack(side="top", padx=10, pady=10) 
# Mode selection (folder vs files)
mode_var = tk.StringVar(value="folder") # default is "folder"
ttk.Label(mode_frame, text="Merging Mode:").pack(side="left", padx=5)
ttk.Radiobutton(mode_frame, text="Use Folder", variable=mode_var, value="folder", command=toggle_mode).pack(side="left", padx=10)
ttk.Radiobutton(mode_frame, text="Use Files", variable=mode_var, value="files", command=toggle_mode).pack(side="left", padx=10)
# header ID
id_label = ttk.Label(mode_frame, text="ID:")
id_var = tk.StringVar(value="Unit name") # column header to identify unique CSV files - default is Unit name
id_entry = ttk.Entry(mode_frame, textvariable=id_var, width=15, justify="left")
id_entry.pack(side="right", padx=5)
id_label.pack(side="right", padx=5)

# Frame for folder
folder_frame = ttk.Frame(left_frame, width=700, height=50, borderwidth=10, relief=tk.GROOVE)
folder_frame.pack_propagate(False) # False - use the defined size
folder_frame.pack(side="top", padx=10, pady=10) 
# Folder selector
ttk.Label(folder_frame, text="Folder:", width=6).pack(side="left")
folder_var = tk.StringVar()
folder_entry = ttk.Entry(folder_frame, textvariable=folder_var, width=52, justify="left")
folder_entry.pack(side="left", padx=5)
folder_browse = tk.Button(folder_frame, text="Browse", width=6, command=browse_folder, font=("Segoe UI", 8))
folder_browse.pack(side="left", padx=5)

# Frame for individual csv files inputs
individual_csv_frame = ttk.Frame(left_frame, width=700, height=350, borderwidth=10, relief=tk.GROOVE)
individual_csv_frame.pack_propagate(False) # False - use the defined size
individual_csv_frame.pack(side="top", padx=10, pady=10) 
# Individual file selectors
entry_vars = [tk.StringVar() for _ in range(10)]
file_entries = []
file_buttons = []

for i in range(10):
    frame = ttk.Frame(individual_csv_frame)
    frame.pack(padx=3, pady=1, fill="x")
    tk.Label(frame, text=f"CSV {i+1:02d}:", width=6, anchor="w").pack(side="left")
    entry = tk.Entry(frame, textvariable=entry_vars[i], width=52, justify="left", state="disabled")
    entry.pack(side="left", padx=5)
    file_entries.append(entry)
    button = tk.Button(frame, text="Browse", width=6, height=1, font=("Segoe UI", 8),
                       state="disabled", command=lambda i=i: browse_file(i))
    button.pack(side="left", padx=1)
    file_buttons.append(button)

# Frame for output folder
output_folder_frame = ttk.Frame(left_frame, width=700, height=50, borderwidth=10, relief=tk.GROOVE)
output_folder_frame.pack_propagate(False) # False - use the defined size
output_folder_frame.pack(side="top", padx=10, pady=10) 
# Output folder
ttk.Label(output_folder_frame, text="Output Folder:", width=12, anchor="w").pack(side="left")
output_folder_var = tk.StringVar()
output_folder_entry = ttk.Entry(output_folder_frame, textvariable=output_folder_var, width=46, justify="left")
output_folder_entry.pack(side="left", padx=10)
tk.Button(output_folder_frame, text="Browse", width=6, height=1, font=("Segoe UI", 8),
          command=browse_output_folder).pack(side="left", padx=1)

# Frame for output filename
output_filename_frame = ttk.Frame(left_frame, width=700, height=50, borderwidth=10, relief=tk.GROOVE)
output_filename_frame.pack_propagate(False) # False - use the defined size
output_filename_frame.pack(side="top", padx=10, pady=10) 
# Output filename + Create button
ttk.Label(output_filename_frame, text="Output name:", width=12, anchor="w").pack(side="left")
output_filename_var = tk.StringVar()
output_filename_entry = ttk.Entry(output_filename_frame, textvariable=output_filename_var, width=35, justify="left")
output_filename_entry.pack(side="left", padx=5)
# button for Create
create_merged_button = tk.Button(output_filename_frame, text="Create", font=("Segoe UI", 10, "bold"),
                          bg="#4CAF50", fg="white", padx=10, pady=2, command=create_merged_file)
create_merged_button.pack(side="left", padx=10)
# button for Update - TODO: create a function for update later
update_merged_button = tk.Button(output_filename_frame, text="Update", font=("Segoe UI", 10, "bold"),
                          bg="#4CAF50", fg="white", padx=10, pady=2, command=update_merged_file)
update_merged_button.pack(side="left", padx=10)

# RIGHT FRAME 
right_frame = tk.Frame(root, width=380, height=650, borderwidth=2, relief=tk.GROOVE)
right_frame.pack_propagate(False) # False - use the defined size
right_frame.pack(side="left", padx=10, pady=10)

# Label
tk.Label(right_frame, text="HEADERS").pack(anchor="w", padx=5)

# Scrollable Listbox for headers
header_frame = ttk.Frame(right_frame)
header_frame.pack(fill="both", expand=True, padx=5)
header_listbox = tk.Listbox(header_frame, selectmode="extended", width=25, height=10)
header_listbox.pack(side="left", fill="both", expand=True, padx=5)
header_scrollbar = tk.Scrollbar(header_frame, orient="vertical", command=header_listbox.yview)
header_scrollbar.pack(side="right", fill="y")
header_listbox.config(yscrollcommand=header_scrollbar.set)

# ---- Axis, Legend, Values ----
# Frame for axis
axis_frame = ttk.Frame(right_frame, width=380, height=50, borderwidth=10, relief=tk.GROOVE)
axis_frame.pack_propagate(False) # False - use the defined size
axis_frame.pack(side="top", padx=5, pady=5) 
ttk.Label(axis_frame, text="Axis:").pack(side='left', padx=5)
axis_var = tk.StringVar()
ttk.Entry(axis_frame, textvariable=axis_var, width=20).pack(side='right', padx=6)

# Frame for legend
legend_frame = ttk.Frame(right_frame, width=380, height=50, borderwidth=10, relief=tk.GROOVE)
legend_frame.pack_propagate(False) # False - use the defined size
legend_frame.pack(side="top", padx=5, pady=5) 
ttk.Label(legend_frame, text="Legend:").pack(side='left', padx=5)
legend_var = tk.StringVar()
ttk.Entry(legend_frame, textvariable=legend_var, width=20).pack(side='right', padx=6)

# Frame for values
value_frame = ttk.Frame(right_frame, width=700, height=50, borderwidth=10, relief=tk.GROOVE)
value_frame.pack_propagate(False) # False - use the defined size
value_frame.pack(side="top", padx=5, pady=5) 
ttk.Label(value_frame, text="Value:").pack(side='left', padx=5)
values_var = tk.StringVar()
ttk.Entry(value_frame, textvariable=values_var, width=20).pack(side='right', padx=6)

# ---- Create Pivot Chart Button ----
# Frame for generate pivot chart 
generate_pc_frame = ttk.Frame(right_frame, width=700, height=50, borderwidth=10, relief=tk.GROOVE)
generate_pc_frame.pack_propagate(False) # False - use the defined size
generate_pc_frame.pack(side="top", padx=5, pady=5) 
# Output pivotchart name + Create button
ttk.Label(generate_pc_frame, text="Chart name:").pack(side="left")
pivotChart_var = tk.StringVar()
pivotChart_entry = ttk.Entry(generate_pc_frame, textvariable=pivotChart_var, width=20, justify="left")
pivotChart_entry.pack(side="right", padx=5)

gen_pivot_button = tk.Button(right_frame, text="Generate PivotChart", width=18, height=1, font=("Segoe UI", 10, "bold"), 
                             bg="#4CAF50",fg="white", command=generate_pivot_chart)
gen_pivot_button.pack(side="top", padx=8, pady=8)

root.mainloop()
