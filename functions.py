import pandas as pd
import glob
import os
from pandas.io.formats import excel
from openpyxl import load_workbook
import win32com.client
import win32gui
import time
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_by_merging(all_files, output_folder, output_filename):   
    # Folder containing all CSV files 
    output_file = os.path.join(output_folder, output_filename + ".xlsx") 
    merged_df = pd.DataFrame()

    # Read and merge CSV files
    df_list = [pd.read_csv(f, parse_dates=False) for f in all_files]
    merged_df = pd.concat(df_list, ignore_index=True)

    # Write to one Excel file
    merged_df.to_excel(output_file, index=False, sheet_name="Raw Data")
    print(f"Merged {len(all_files)} files into {output_file}")  
    
    # Convert to Table
    convert_to_RawData_table(output_file)

def convert_to_RawData_table(merged_file):
    # Convert the Raw Data to Table to easily refresh pivot chart later
    wb = load_workbook(merged_file)
    ws = wb["Raw Data"]
    end_row, end_col = ws.max_row, ws.max_column
    table_range = f"A1:{ws.cell(end_row, end_col).coordinate}"

    # clear old tables because it is invalid - range have changed
    if ws._tables:
        for tbl in list(ws._tables.values()):
            ws._tables.pop(tbl.name)

    table = Table(displayName="RawDataTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleLight1")
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(merged_file)
    
def generate_pivotChart(merged_path, pc_axis, pc_legend, pc_values, pivotName):
    # Paths to your files
    updater_path = r"C:\Users\Harvey\Desktop\Projects\csv_to_pivotChart\Updater.xlsm"

    # Start Excel
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # Set to True to automatically open the merged file after macro runs

    # Open the Updater file
    updater_wb = excel.Workbooks.Open(updater_path)

    # Run your macro
    # excel.Application.Run("Module2.CreatePivotChartInMergedData")
    # pc_axis = "DateTime"
    # pc_legend = "Unit name"
    # pc_values = "Voltage"
    # pivotName = pc_axis + " over " + pc_values

    excel.Application.Run("Module3.CreatePivotChartInMergedDataVariable", merged_path, pivotName, pc_axis, pc_legend, pc_values)

    # Save & close Updater (optional)
    updater_wb.Save()
    updater_wb.Close()

    # Open the merged data file
    merged_wb = excel.Workbooks.Open(merged_path)
    print(f"Opened merged file: {merged_wb.Name}")

    # (Optional) — you can activate a specific sheet like PivotChart:
    merged_wb.Sheets(pivotName).Activate()

    # Give Excel some time to fully open
    time.sleep(0.5)

    # Bring Excel to the front
    win32gui.SetForegroundWindow(win32gui.FindWindow(None, excel.Caption))

def rebindPivotSources(merged_file):
     # Paths to your files
    updater_path = r"C:\Users\Harvey\Desktop\Projects\csv_to_pivotChart\Updater.xlsm"

    # Start Excel
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # Set to True to automatically open the merged file after macro runs

    # Open the Updater file
    updater_wb = excel.Workbooks.Open(updater_path)

    # Run the rebinding macro
    excel.Application.Run("Module4.RebindPivotSources", merged_file)

    # Save & close Updater (optional)
    updater_wb.Save()
    updater_wb.Close()

    # Open the merged data file
    merged_wb = excel.Workbooks.Open(merged_file)
    print(f"Opened merged file: {merged_wb.Name}")

    # Give Excel some time to fully open
    time.sleep(0.5)

    # Bring Excel to the front
    win32gui.SetForegroundWindow(win32gui.FindWindow(None, excel.Caption))
